import openpyxl


class ExcelRead():

    filepath = "C:\\Users\\DELL\\PycharmProjects\\PythonProject1\\Input\\Swiggy.xlsx"

    dict ={}
    def Readdata(self,SheetName):

        self.dict.clear()
        # open the excel file
        workbook = openpyxl.load_workbook(self.filepath)
        # move to the specific sheet
        Sheet = workbook[SheetName]
        # go to the specific row and column
        column = Sheet.cell(row=2, column=1).value
        # print(column)
        # to get a Max used row
        Total_Rows = Sheet.max_row
        # to get a max column
        Total_column = Sheet.max_column
        # to navigate in to all the rows
        for eachrow in range(2, Total_Rows + 1):
            # to navigate in to all the columns
            for eachcolumn in range(1, Total_column + 1):
                # get the value from each row and column
                eachCell = Sheet.cell(row=eachrow, column=eachcolumn).value
                # print(eachCell)
                # key =Sheet.cell(row=1, column=eachcolumn).value +str(eachrow)
                # key = Sheet.cell(row=1, column=eachcolumn).value + Sheet.cell(row=eachrow, column=1).value
                key = Sheet.cell(row=1, column=eachcolumn).value + str(eachrow)

                self.dict[key] = eachCell
        print("*********Dictionalty values**********")
        print(self.dict)
        workbook.close()
        # for eachvalue in self.dict.values():
        # print(eachvalue)
        #return self.dict

        # print(self.dict["FeesPaidFT2"])

obj = ExcelRead()
obj.Readdata("Hotels")
obj.Readdata("A2BItem")

