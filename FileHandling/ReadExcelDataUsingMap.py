import openpyxl


class ReadExcelDataUsingMap():

    filepath = "C:\\Users\\DELL\\PycharmProjects\\PythonProject1\\Input\\Swiggy.xlsx"

    def REaddatawithMap(self,SheetName):
        workbook = openpyxl.load_workbook(self.filepath)
        # move to the specific sheet
        Sheet = workbook[SheetName]
        for row in Sheet.iter_rows(min_row=1, values_only=True):  # Skip header
            row
            print(row)
            #print(f"Name: {name}, Date: {date}")

        print(type(row))

        #map(Function,iteratabledata(list,tuple,set)
        names = list(map(lambda row: row[2].value, Sheet.iter_rows(min_row=2)))
        print("list of name : ",names)

        for i, row in enumerate(Sheet.iter_rows(values_only=True), start=1):
            print(f"Row {i}: {row}")

        #filtered_data = list(filter(lambda row1: row1['Status'] != "InActive", names))
        #print(filtered_data)

        workbook.close()

        data = [
            {'Name': 'Alice', 'Status': 'Active'},
            {'Name': 'Bob', 'Status': 'Inactive'},
            {'Name': 'Clara', 'Status': 'Active'}
        ]
        active_rows = list(filter(lambda row: row['Status'] == 'Active', data))
        print(active_rows)

    def readDataList(self,SheetName):
        workbook = openpyxl.load_workbook(self.filepath)
        # move to the specific sheet
        ws = workbook[SheetName]
        # Step 1: Read column headers from the first row
        headers = [cell.value for cell in ws[1]]  # e.g., ['Name', 'Age', 'City']

        # Step 2: Iterate through each remaining row and convert to dictionary
        data_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            row_dict = dict(zip(headers, row))  # pair header with each cell value
            data_list.append(row_dict)

        # Now data_list is a list of dictionaries
        for entry in data_list:
            print(entry)

        print(data_list)
        active_rows = list(filter(lambda row: row['Status'] == 'Active', data_list))
        print(active_rows)
        workbook.close()

obj =ReadExcelDataUsingMap()
obj.readDataList("Hotels")