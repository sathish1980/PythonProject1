import openpyxl


class ExcelWrite():

    Ofilepath = "C:\\Users\\DELL\\PycharmProjects\\PythonProject1\\Input\\Output.xlsx"
    Rfilepath = "C:\\Users\\DELL\\PycharmProjects\\PythonProject1\\Input\\Swiggy.xlsx"
    dict ={}
    def Writedata(self):
        # open the excel file
        writeworkbook = openpyxl.Workbook()
        # move to the specific sheet
        sheet = writeworkbook.active
        sheet.title = "output"
        # go to the specific row and column
        # c= sheet.cell(row = 1, column = 1)
        #
        sheet.cell(row=1, column=1).value = "Sathish"
        sheet.cell(row=1, column=2).value = "kumar"
        sheet.cell(row=1, column=3).value = "R"
        sheet.cell(row=2, column=1).value = "raja"
        sheet.cell(row=2, column=2).value = "kumar"
        sheet.cell(row=2, column=3).value = "M"

        writeworkbook.save(self.filepath)

        print("Done")

    def ReadandWrite(self,SheetName):

        #
        # open the excel file
        workbook = openpyxl.load_workbook(self.Rfilepath)
        # open the excel file
        writeworkbook = openpyxl.Workbook()

        # move to the specific sheet
        Sheet = workbook[SheetName]
        # create asheet
        writesheet = writeworkbook.active
        writesheet.title = "output"
        # to get a Max used row
        Total_Rows = Sheet.max_row
        # to get a max column
        Total_column = Sheet.max_column
        # to navigate in to all the rows
        for eachrow in range(1, Total_Rows + 1):
            # to navigate in to all the columns
            check = Sheet.cell(row=eachrow, column=3).value
            if (check!="InActive"):
                for eachcolumn in range(1, Total_column + 1):
                    # get the value from each row and column
                    eachCell = Sheet.cell(row=eachrow, column=eachcolumn).value
                    writesheet.cell(row=eachrow, column=eachcolumn).value = eachCell

        writeworkbook.save(self.Ofilepath)
        writeworkbook.close()
        print("done")


obj = ExcelWrite()
obj.ReadandWrite("Hotels")

